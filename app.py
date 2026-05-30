from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for
import json, os, io, sqlite3, hashlib, secrets
from functools import wraps

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

DB_FILE = "edubank.db"

DAFTAR_BULAN = {
    "januari":31,"februari":28,"maret":31,"april":30,"mei":31,"juni":30,
    "juli":31,"agustus":31,"september":30,"oktober":31,"november":30,"desember":31
}

# ============================================================
# DATABASE SETUP
# ============================================================
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # Tabel users
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')
    # Tabel data keuangan per user (simpan sebagai JSON blob)
    c.execute('''
        CREATE TABLE IF NOT EXISTS user_data (
            user_id INTEGER PRIMARY KEY,
            bulan_aktif TEXT DEFAULT '',
            jumlah_hari INTEGER DEFAULT 0,
            data_keuangan TEXT DEFAULT '{}',
            data_ewallet TEXT DEFAULT '{}',
            data_tabungan TEXT DEFAULT '{"saldo": 0, "riwayat": []}',
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    conn.commit()
    conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def get_user_by_email(email):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE email = ?", (email,))
    user = c.fetchone()
    conn.close()
    return user

def get_user_data(user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT * FROM user_data WHERE user_id = ?", (user_id,))
    row = c.fetchone()
    conn.close()
    if row:
        return {
            "bulan_aktif": row[1],
            "jumlah_hari": row[2],
            "data_keuangan": json.loads(row[3]),
            "data_ewallet": json.loads(row[4]),
            "data_tabungan": json.loads(row[5])
        }
    return {
        "bulan_aktif": "",
        "jumlah_hari": 0,
        "data_keuangan": {},
        "data_ewallet": {},
        "data_tabungan": {"saldo": 0, "riwayat": []}
    }

def save_user_data(user_id, data):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        INSERT INTO user_data (user_id, bulan_aktif, jumlah_hari, data_keuangan, data_ewallet, data_tabungan)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            bulan_aktif   = excluded.bulan_aktif,
            jumlah_hari   = excluded.jumlah_hari,
            data_keuangan = excluded.data_keuangan,
            data_ewallet  = excluded.data_ewallet,
            data_tabungan = excluded.data_tabungan
    ''', (
        user_id,
        data.get("bulan_aktif", ""),
        data.get("jumlah_hari", 0),
        json.dumps(data.get("data_keuangan", {})),
        json.dumps(data.get("data_ewallet", {})),
        json.dumps(data.get("data_tabungan", {"saldo": 0, "riwayat": []}))
    ))
    conn.commit()
    conn.close()

# ============================================================
# AUTH DECORATOR
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

# ============================================================
# ROUTES — AUTH
# ============================================================
@app.route("/")
def root():
    if "user_id" in session:
        return redirect(url_for("index"))
    return redirect(url_for("login_page"))

@app.route("/login")
def login_page():
    if "user_id" in session:
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/api/register", methods=["POST"])
def api_register():
    data  = request.get_json()
    nama  = data.get("nama", "").strip()
    email = data.get("email", "").strip().lower()
    pw    = data.get("password", "")

    if not nama or not email or not pw:
        return jsonify({"status": "error", "message": "Semua field wajib diisi!"}), 400
    if len(pw) < 6:
        return jsonify({"status": "error", "message": "Password minimal 6 karakter!"}), 400
    if get_user_by_email(email):
        return jsonify({"status": "error", "message": "Email sudah terdaftar!"}), 400

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT INTO users (nama, email, password) VALUES (?, ?, ?)",
              (nama, email, hash_password(pw)))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "message": "Registrasi berhasil! Silakan login."})

@app.route("/api/login", methods=["POST"])
def api_login():
    data  = request.get_json()
    email = data.get("email", "").strip().lower()
    pw    = data.get("password", "")

    user = get_user_by_email(email)
    if not user or user[3] != hash_password(pw):
        return jsonify({"status": "error", "message": "Email atau password salah!"}), 401

    session["user_id"] = user[0]
    session["nama"]    = user[1]
    session["email"]   = user[2]
    return jsonify({"status": "ok", "nama": user[1], "email": user[2]})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"status": "ok"})

@app.route("/api/me")
@login_required
def api_me():
    return jsonify({"nama": session["nama"], "email": session["email"]})

# ============================================================
# ROUTES — APP
# ============================================================
@app.route("/app")
@login_required
def index():
    return render_template("index.html", nama=session["nama"])

@app.route("/api/state")
@login_required
def api_state():
    return jsonify(get_user_data(session["user_id"]))

@app.route("/api/save", methods=["POST"])
@login_required
def api_save():
    d = request.get_json()
    save_user_data(session["user_id"], d)
    return jsonify({"status": "ok"})

# ── Export Excel ──
@app.route("/api/export-excel")
@login_required
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    d = get_user_data(session["user_id"])
    data_keuangan = d.get("data_keuangan", {})
    data_ewallet  = d.get("data_ewallet", {})
    data_tabungan = d.get("data_tabungan", {"saldo": 0, "riwayat": []})

    def hf(c): return PatternFill("solid", fgColor=c)
    def tb():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def sh(cell, bg="2C3E50", fg="FFFFFF"):
        cell.fill = hf(bg); cell.font = Font(bold=True, color=fg, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = tb()
    def sc(cell, align="left", num_fmt=None, bg=None):
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center"); cell.border = tb()
        if num_fmt: cell.number_format = num_fmt
        if bg: cell.fill = hf(bg)

    wb = Workbook()
    first = True
    kc = {"kebutuhan": "D5F5E3", "impulsif": "FADBD8", "tabungan": "D6EAF8"}

    for bulan, data in data_keuangan.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = bulan[:15]; first = False
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 18

        ws["A1"] = f"LAPORAN KEUANGAN — {bulan.upper()} ({session['nama']})"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = hf("1A5276")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:D1")

        ws["A3"] = "PEMASUKAN (SAKU)"; ws["A3"].font = Font(bold=True, color="FFFFFF")
        ws["A3"].fill = hf("27AE60"); ws.merge_cells("A3:D3")
        for ci, h in enumerate(["Hari","Keterangan","Jumlah (Rp)"], 1):
            sh(ws.cell(4, ci, h), bg="1E8449")
        ri = 5; total_masuk = 0
        for hari, items in data.get("pemasukan", {}).items():
            for item in items:
                sc(ws.cell(ri,1,hari), align="center")
                sc(ws.cell(ri,2,item["keterangan"]))
                sc(ws.cell(ri,3,item["jumlah"]), align="right", num_fmt="#,##0", bg="E8F8F5")
                total_masuk += item["jumlah"]; ri += 1
        sc(ws.cell(ri,2,"TOTAL PEMASUKAN"), align="right")
        sc(ws.cell(ri,3,total_masuk), align="right", num_fmt="#,##0", bg="D5F5E3")
        ws.cell(ri,2).font = Font(bold=True); ri += 2

        ws.cell(ri,1,"PENGELUARAN (SAKU)").font = Font(bold=True, color="FFFFFF")
        ws.cell(ri,1).fill = hf("E74C3C"); ws.merge_cells(f"A{ri}:D{ri}"); ri += 1
        for ci, h in enumerate(["Hari","Keterangan","Kategori","Jumlah (Rp)"], 1):
            sh(ws.cell(ri, ci, h), bg="922B21"); ri += 1
        total_keluar = 0
        for hari, items in data.get("pengeluaran", {}).items():
            for item in items:
                kat = item.get("kategori","kebutuhan")
                sc(ws.cell(ri,1,hari), align="center")
                sc(ws.cell(ri,2,item["keterangan"]))
                sc(ws.cell(ri,3,kat.capitalize()), align="center", bg=kc.get(kat,"FFFFFF"))
                sc(ws.cell(ri,4,item["jumlah"]), align="right", num_fmt="#,##0", bg="FADBD8")
                total_keluar += item["jumlah"]; ri += 1
        sc(ws.cell(ri,3,"TOTAL PENGELUARAN"), align="right")
        sc(ws.cell(ri,4,total_keluar), align="right", num_fmt="#,##0", bg="FADBD8")
        ws.cell(ri,3).font = Font(bold=True); ri += 2

        saldo_awal  = data.get("saldo_awal") or 0
        saldo_akhir = saldo_awal + total_masuk - total_keluar
        total_ew    = sum(info.get("saldo",0) for info in data_ewallet.values())

        ws.cell(ri,1,"RINGKASAN").font = Font(bold=True, color="FFFFFF")
        ws.cell(ri,1).fill = hf("8E44AD"); ws.merge_cells(f"A{ri}:D{ri}"); ri += 1
        for label, val in [
            ("Saldo Awal (Saku)", saldo_awal),
            ("Pemasukan Saku",    total_masuk),
            ("Pengeluaran Saku",  total_keluar),
            ("Saldo Saku Akhir",  saldo_akhir),
            ("Total E-Wallet",    total_ew),
            ("SALDO TOTAL",       saldo_akhir + total_ew),
            ("Saldo Tabungan",    data_tabungan.get("saldo",0)),
        ]:
            sc(ws.cell(ri,1,label))
            sc(ws.cell(ri,2,val), align="right", num_fmt="#,##0",
               bg="D5F5E3" if val>=0 else "FADBD8")
            if label in ("SALDO TOTAL","Saldo Saku Akhir"):
                ws.cell(ri,1).font = Font(bold=True)
                ws.cell(ri,2).font = Font(bold=True)
            ri += 1

    ws2 = wb.create_sheet("Ringkasan Semua Bulan")
    ws2.sheet_view.showGridLines = False
    hdrs = ["Bulan","Saldo Awal","Pemasukan","Pengeluaran","Saldo Saku Akhir","Total E-Wallet","Saldo Total","Status"]
    widths = [14,18,18,18,20,18,18,14]
    for col,(h,w) in enumerate(zip(hdrs,widths),1):
        sh(ws2.cell(1,col,h), bg="1A5276")
        ws2.column_dimensions[get_column_letter(col)].width = w

    total_ew_all = sum(info.get("saldo",0) for info in data_ewallet.values())
    n_bulan = 0
    for ri2,(bulan,data) in enumerate(data_keuangan.items(), 2):
        saldo_awal  = data.get("saldo_awal") or 0
        masuk  = sum(i["jumlah"] for h in data.get("pemasukan",{}).values()  for i in h)
        keluar = sum(i["jumlah"] for h in data.get("pengeluaran",{}).values() for i in h)
        akhir  = saldo_awal + masuk - keluar
        total  = akhir + total_ew_all
        pct    = (akhir/saldo_awal*100) if saldo_awal>0 else 0
        status = "Aman" if pct>=70 else "Waspada" if pct>=40 else "Hampir Habis" if pct>=10 else "Defisit"
        bg_st  = "D5F5E3" if status=="Aman" else "FDEBD0" if status=="Waspada" else "FADBD8"
        for ci,val in enumerate([bulan,saldo_awal,masuk,keluar,akhir,total_ew_all,total,status],1):
            sc(ws2.cell(ri2,ci,val), align="center",
               num_fmt="#,##0" if ci in [2,3,4,5,6,7] else None,
               bg=bg_st if ci==8 else ("F2F3F4" if ri2%2==0 else "FFFFFF"))
        n_bulan += 1

    if n_bulan > 0:
        chart = BarChart(); chart.type = "col"; chart.style = 10
        chart.width = 22; chart.height = 14; chart.title = "Pemasukan vs Pengeluaran"
        chart.add_data(Reference(ws2,min_col=3,max_col=4,min_row=1,max_row=n_bulan+1), titles_from_data=True)
        chart.set_categories(Reference(ws2,min_col=1,min_row=2,max_row=n_bulan+1))
        chart.series[0].graphicalProperties.solidFill = "27AE60"
        chart.series[1].graphicalProperties.solidFill = "E74C3C"
        ws2.add_chart(chart, "J2")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f"laporan_edubank_{session['nama']}.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    init_db()
    app.run(debug=True)
