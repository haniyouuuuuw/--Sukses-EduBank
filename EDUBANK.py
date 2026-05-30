import json
import os

# ===== DATA GLOBAL =====
# data_keuangan: { "Januari": { "saldo_awal": 0, "pemasukan": {}, "pengeluaran": {} } }
# data_ewallet:  { "GoPay (Fina)": { "saldo": 0, "pengguna": "", "transaksi": [] } }
# data_tabungan: { "saldo": 0, "riwayat": [ { "jenis": "tambah"/"kurang", "jumlah": 0, "keterangan": "" } ] }

data_keuangan = {}
data_ewallet  = {}
data_tabungan = {"saldo": 0, "riwayat": []}
bulan_aktif   = ""
jumlah_hari   = 0

DAFTAR_BULAN = {
    "januari": 31, "februari": 28, "maret": 31,
    "april": 30,   "mei": 31,      "juni": 30,
    "juli": 31,    "agustus": 31,  "september": 30,
    "oktober": 31, "november": 30, "desember": 31
}
URUTAN_BULAN = list(DAFTAR_BULAN.keys())

# ============================================================
# SIMPAN & LOAD
# ============================================================
def simpan_data():
    export = {
        "bulan_aktif":   bulan_aktif,
        "jumlah_hari":   jumlah_hari,
        "data_keuangan": data_keuangan,
        "data_ewallet":  data_ewallet,
        "data_tabungan": data_tabungan
    }
    with open("data.json", "w") as f:
        json.dump(export, f, indent=2)
    print("✅ Data berhasil disimpan!")

def load_data():
    global bulan_aktif, jumlah_hari, data_keuangan, data_ewallet, data_tabungan
    if os.path.exists("data.json"):
        with open("data.json") as f:
            d = json.load(f)
        bulan_aktif   = d.get("bulan_aktif", "")
        jumlah_hari   = d.get("jumlah_hari", 0)
        data_keuangan = d.get("data_keuangan", {})
        data_ewallet  = d.get("data_ewallet", {})
        data_tabungan = d.get("data_tabungan", {"saldo": 0, "riwayat": []})
        print("✅ Data sebelumnya berhasil dimuat!")
    else:
        print("ℹ️  Mulai dari awal.")

# ============================================================
# VALIDASI
# ============================================================
def cek_bulan():
    if not bulan_aktif:
        print("\n⚠️  Pilih bulan dulu! (Menu 1)")
        return False
    return True

def cek_saldo():
    if not bulan_aktif:
        print("\n⚠️  Pilih bulan dulu! (Menu 1)")
        return False
    if data_keuangan[bulan_aktif].get("saldo_awal") is None:
        print("\n⚠️  Input saldo awal dulu! (Menu 2)")
        return False
    return True

# ============================================================
# HELPER: hitung saldo akhir suatu bulan
# ============================================================
def hitung_saldo_akhir(bulan):
    d = data_keuangan.get(bulan, {})
    saldo  = d.get("saldo_awal") or 0
    masuk  = sum(i["jumlah"] for h in d.get("pemasukan", {}).values()  for i in h)
    keluar = sum(i["jumlah"] for h in d.get("pengeluaran", {}).values() for i in h)
    return saldo + masuk - keluar

def hitung_saldo_akhir_ew(nama_ew, sampai_bulan=None):
    """Hitung saldo e-wallet berdasarkan transaksi (untuk carry-over)."""
    info = data_ewallet.get(nama_ew, {})
    return info.get("saldo", 0)

def bulan_sebelumnya(bulan):
    idx = URUTAN_BULAN.index(bulan.lower())
    if idx == 0:
        return None
    return URUTAN_BULAN[idx - 1].capitalize()

# ============================================================
# MENU 1 — PILIH BULAN
# ============================================================
def pilih_bulan():
    global bulan_aktif, jumlah_hari
    print("\n=== PILIH BULAN ===")
    for i, b in enumerate(DAFTAR_BULAN, 1):
        print(f"  {i:2}. {b.capitalize()}")

    bulan = input("\nKetik nama bulan (contoh: januari): ").strip().lower()
    if bulan not in DAFTAR_BULAN:
        print("❌ Bulan tidak valid!")
        return

    bulan_aktif = bulan.capitalize()
    jumlah_hari = DAFTAR_BULAN[bulan]

    if bulan_aktif not in data_keuangan:
        # Cek apakah ada bulan sebelumnya untuk carry-over
        prev = bulan_sebelumnya(bulan_aktif)
        saldo_carry = 0
        ew_carry    = {}

        if prev and prev in data_keuangan:
            saldo_carry = hitung_saldo_akhir(prev)
            # Carry saldo e-wallet
            for nama, info in data_ewallet.items():
                ew_carry[nama] = info.get("saldo", 0)

            print(f"\n  Ada data bulan {prev}.")
            print(f"  Saldo akhir bulan {prev}: Rp {saldo_carry:,}")
            print("  Carry-over saldo ke bulan ini? (y = lanjut, n = mulai dari 0)")
            jawab = input("  Pilihan: ").strip().lower()
            if jawab != "y":
                saldo_carry = 0
                ew_carry    = {}
                print("  Saldo dimulai dari Rp 0.")

        data_keuangan[bulan_aktif] = {
            "saldo_awal":  saldo_carry,
            "pemasukan":   {},
            "pengeluaran": {}
        }
        # Update saldo e-wallet carry-over jika ada
        if ew_carry:
            for nama, saldo in ew_carry.items():
                if nama in data_ewallet:
                    # Simpan saldo awal bulan baru tapi jangan override transaksi
                    pass  # saldo sudah tersimpan di data_ewallet[nama]["saldo"]

        print(f"✅ Bulan {bulan_aktif} dibuat ({jumlah_hari} hari), saldo awal: Rp {saldo_carry:,}")
    else:
        saldo = data_keuangan[bulan_aktif].get("saldo_awal")
        print(f"✅ Bulan {bulan_aktif} dipilih ({jumlah_hari} hari)")
        if saldo is not None:
            print(f"   Saldo awal bulan ini: Rp {saldo:,}")
            print(f"   Saldo akhir estimasi: Rp {hitung_saldo_akhir(bulan_aktif):,}")

# ============================================================
# MENU 2 — INPUT / EDIT SALDO AWAL
# ============================================================
def input_saldo():
    if not cek_bulan():
        return
    print(f"\n=== INPUT SALDO AWAL — {bulan_aktif} ===")
    saldo_lama = data_keuangan[bulan_aktif].get("saldo_awal")
    if saldo_lama is not None:
        print(f"Saldo awal sekarang: Rp {saldo_lama:,}")
    try:
        saldo = int(input("Masukkan saldo awal bulan ini (saldo keuangan umum/saku): Rp "))
        data_keuangan[bulan_aktif]["saldo_awal"] = saldo
        print(f"✅ Saldo awal {bulan_aktif}: Rp {saldo:,} berhasil disimpan.")
    except:
        print("❌ Input harus angka!")

# ============================================================
# MENU 3 — TAMBAH TRANSAKSI SAKU
# ============================================================
def tambah_transaksi():
    if not cek_saldo():
        return

    print(f"\n=== TAMBAH TRANSAKSI SAKU — {bulan_aktif} ===")
    print("(Transaksi ini mempengaruhi Saldo Keuangan Umum/Saku)")

    try:
        hari = int(input(f"Hari ke berapa? (1-{jumlah_hari}): "))
        if hari < 1 or hari > jumlah_hari:
            print("❌ Hari tidak valid!"); return
    except:
        print("❌ Input harus angka!"); return

    print("\nJenis transaksi:")
    print("  a = Pemasukan  (uang masuk: gaji, uang jajan, hadiah)")
    print("  b = Pengeluaran (uang keluar: beli sesuatu, bayar tagihan)")
    jenis = input("Pilih (a/b): ").strip().lower()
    if jenis not in ["a", "b"]:
        print("❌ Pilihan tidak valid!"); return

    keterangan = input("Keterangan: ").strip()
    if not keterangan:
        print("❌ Keterangan tidak boleh kosong!"); return

    try:
        jumlah = int(input("Jumlah: Rp "))
        if jumlah <= 0:
            print("❌ Jumlah harus lebih dari 0!"); return
    except:
        print("❌ Input harus angka!"); return

    key = f"Day {hari}"

    if jenis == "a":
        data_keuangan[bulan_aktif]["pemasukan"].setdefault(key, []).append({
            "keterangan": keterangan,
            "jumlah":     jumlah
        })
        print(f"✅ Pemasukan Rp {jumlah:,} ({keterangan}) berhasil ditambahkan ke saldo umum.")

    else:
        print("\nKategori pengeluaran:")
        print("  1. Kebutuhan  (makan, transport, bayar tagihan, kesehatan)")
        print("  2. Impulsif   (flash sale, jajan random, hiburan mendadak)")
        print("  3. Tabungan   (label saja — tidak otomatis masuk menu Tabungan)")
        try:
            kat = int(input("Pilih kategori (1/2/3): "))
            kat_map = {1: "kebutuhan", 2: "impulsif", 3: "tabungan"}
            if kat not in kat_map:
                print("❌ Kategori tidak valid!"); return
            kategori = kat_map[kat]
        except:
            print("❌ Input harus angka!"); return

        data_keuangan[bulan_aktif]["pengeluaran"].setdefault(key, []).append({
            "keterangan": keterangan,
            "kategori":   kategori,
            "jumlah":     jumlah
        })
        print(f"✅ Pengeluaran Rp {jumlah:,} [{kategori}] ({keterangan}) berhasil ditambahkan.")

# ============================================================
# MENU 4 — RINGKASAN BULAN INI
# ============================================================
def ringkasan():
    if not cek_saldo():
        return

    d      = data_keuangan[bulan_aktif]
    saldo  = d["saldo_awal"]
    masuk  = sum(i["jumlah"] for h in d["pemasukan"].values()  for i in h)
    keluar = sum(i["jumlah"] for h in d["pengeluaran"].values() for i in h)
    akhir  = saldo + masuk - keluar

    # Total saldo e-wallet bulan ini
    total_ew = sum(info.get("saldo", 0) for info in data_ewallet.values())
    saldo_total = akhir + total_ew

    kat_total = {"kebutuhan": 0, "impulsif": 0, "tabungan": 0}
    for h in d["pengeluaran"].values():
        for item in h:
            kat = item.get("kategori", "kebutuhan")
            kat_total[kat] += item["jumlah"]

    # Tambah kategori dari e-wallet
    for info in data_ewallet.values():
        for t in info.get("transaksi", []):
            if t.get("jenis") == "pengeluaran":
                kat = t.get("kategori", "kebutuhan")
                if kat in kat_total:
                    kat_total[kat] += t["jumlah"]

    print(f"\n{'='*42}")
    print(f"  RINGKASAN {bulan_aktif.upper()}")
    print(f"{'='*42}")
    print(f"  Saldo Awal (Saku) : Rp {saldo:>12,}")
    print(f"  Pemasukan Saku    : Rp {masuk:>12,}")
    print(f"  Pengeluaran Saku  : Rp {keluar:>12,}")
    print(f"  {'─'*36}")
    print(f"  Saldo Saku Akhir  : Rp {akhir:>12,}")
    print(f"  Total Saldo E-Wallet: Rp {total_ew:>10,}")
    print(f"  {'─'*36}")
    print(f"  SALDO TOTAL       : Rp {saldo_total:>12,}")
    print(f"  Saldo Tabungan    : Rp {data_tabungan['saldo']:>12,}")
    print(f"\n  Rincian Pengeluaran (Saku + E-Wallet):")
    print(f"    Kebutuhan       : Rp {kat_total['kebutuhan']:>10,}")
    print(f"    Impulsif        : Rp {kat_total['impulsif']:>10,}")
    print(f"    Tabungan (label): Rp {kat_total['tabungan']:>10,}")

    total_keluar_all = sum(kat_total.values())
    if saldo > 0:
        pct   = (akhir / saldo) * 100
        impct = (kat_total["impulsif"] / total_keluar_all * 100) if total_keluar_all > 0 else 0
        if pct >= 70:   status = "🟢 Aman"
        elif pct >= 40: status = "🟡 Waspada"
        elif pct >= 10: status = "🟠 Hampir Habis"
        elif pct >= 0:  status = "🔴 Kritis"
        else:           status = "❌ Defisit"
        print(f"\n  Sisa Saldo Saku : {pct:.1f}%")
        print(f"  Status          : {status}")
        if impct >= 50:
            print(f"\n  ⚠️  {impct:.0f}% pengeluaranmu impulsif! Hati-hati.")
        elif impct >= 30:
            print(f"\n  ⚠️  {impct:.0f}% pengeluaran impulsif. Mulai dikurangi ya.")
    print(f"{'='*42}")

# ============================================================
# MENU 5 — TOTAL PER HARI
# ============================================================
def total_per_day():
    if not cek_bulan():
        return
    d = data_keuangan[bulan_aktif]
    print(f"\n=== TOTAL PER HARI — {bulan_aktif} ===")
    ada = False
    for i in range(1, jumlah_hari + 1):
        key    = f"Day {i}"
        masuk  = sum(x["jumlah"] for x in d["pemasukan"].get(key, []))
        keluar = sum(x["jumlah"] for x in d["pengeluaran"].get(key, []))
        if masuk or keluar:
            print(f"  {key:<8} | Masuk: Rp {masuk:>9,} | Keluar: Rp {keluar:>9,}")
            ada = True
    if not ada:
        print("  Belum ada transaksi bulan ini.")

# ============================================================
# MENU 6 — EDIT / HAPUS TRANSAKSI
# ============================================================
def edit_hapus():
    if not data_keuangan:
        print("\n⚠️  Belum ada data!"); return

    print("\n=== EDIT / HAPUS TRANSAKSI ===")
    bulan_list = list(data_keuangan.keys())
    for i, b in enumerate(bulan_list, 1):
        print(f"  {i}. {b}")
    try:
        idx   = int(input("Pilih nomor bulan: ")) - 1
        bulan = bulan_list[idx]
    except:
        print("❌ Input tidak valid!"); return

    print(f"\nBulan: {bulan}")
    print("  a = Pemasukan | b = Pengeluaran")
    jenis = input("Pilih (a/b): ").strip().lower()
    if jenis == "a":   kategori_key = "pemasukan"
    elif jenis == "b": kategori_key = "pengeluaran"
    else:
        print("❌ Tidak valid!"); return

    d        = data_keuangan[bulan]
    all_data = [(h, item) for h in d[kategori_key] for item in d[kategori_key][h]]
    if not all_data:
        print("Tidak ada data."); return

    for i, (h, item) in enumerate(all_data):
        kat = f" [{item.get('kategori','')}]" if kategori_key == "pengeluaran" else ""
        print(f"  {i}. {h} | {item['keterangan']}{kat} | Rp {item['jumlah']:,}")

    try:
        i    = int(input("Pilih nomor data: "))
        hari, item = all_data[i]
    except:
        print("❌ Input tidak valid!"); return

    aksi = input("e = edit jumlah | h = hapus: ").strip().lower()
    if aksi == "e":
        try:
            item["jumlah"] = int(input("Jumlah baru: Rp "))
            print("✅ Data diupdate.")
        except:
            print("❌ Harus angka!")
    elif aksi == "h":
        d[kategori_key][hari].remove(item)
        print("✅ Data dihapus.")
    else:
        print("❌ Tidak valid!")

# ============================================================
# MENU 7 — EXPORT EXCEL (per bulan, beda tabel)
# ============================================================
def export_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    def hf(c): return PatternFill("solid", fgColor=c)
    def tb():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def sh(cell, bg="2C3E50", fg="FFFFFF"):
        cell.fill = hf(bg); cell.font = Font(bold=True, color=fg, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = tb()
    def sc(cell, align="left", num_fmt=None, bg=None):
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center"); cell.border = tb()
        if num_fmt: cell.number_format = num_fmt
        if bg: cell.fill = hf(bg)

    wb = Workbook()
    first = True

    # ── Sheet per bulan ──
    for bulan, data in data_keuangan.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = bulan[:15]
        first = False
        ws.sheet_view.showGridLines = False

        # Header info bulan
        ws["A1"] = f"LAPORAN KEUANGAN — {bulan.upper()}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = hf("1A5276")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:F1")

        # Tabel Pemasukan
        ws["A3"] = "PEMASUKAN (SAKU)"
        ws["A3"].font = Font(bold=True, color="FFFFFF")
        ws["A3"].fill = hf("27AE60")
        ws.merge_cells("A3:F3")

        hdrs = ["Hari", "Keterangan", "Jumlah (Rp)"]
        for ci, h in enumerate(hdrs, 1):
            sh(ws.cell(4, ci, h), bg="1E8449")
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 18

        ri = 5
        total_masuk = 0
        for hari, items in data.get("pemasukan", {}).items():
            for item in items:
                sc(ws.cell(ri, 1, hari), align="center")
                sc(ws.cell(ri, 2, item["keterangan"]))
                sc(ws.cell(ri, 3, item["jumlah"]), align="right", num_fmt="#,##0", bg="E8F8F5")
                total_masuk += item["jumlah"]
                ri += 1
        sc(ws.cell(ri, 2, "TOTAL PEMASUKAN"), align="right")
        sc(ws.cell(ri, 3, total_masuk), align="right", num_fmt="#,##0", bg="D5F5E3")
        ws.cell(ri, 2).font = Font(bold=True)
        ri += 2

        # Tabel Pengeluaran
        ws.cell(ri, 1, "PENGELUARAN (SAKU)").font = Font(bold=True, color="FFFFFF")
        ws.cell(ri, 1).fill = hf("E74C3C")
        ws.merge_cells(f"A{ri}:F{ri}")
        ri += 1

        hdrs2 = ["Hari", "Keterangan", "Kategori", "Jumlah (Rp)"]
        for ci, h in enumerate(hdrs2, 1):
            sh(ws.cell(ri, ci, h), bg="922B21")
        ri += 1

        total_keluar = 0
        kat_total = {"kebutuhan": 0, "impulsif": 0, "tabungan": 0}
        kc = {"kebutuhan": "D5F5E3", "impulsif": "FADBD8", "tabungan": "D6EAF8"}
        for hari, items in data.get("pengeluaran", {}).items():
            for item in items:
                kat = item.get("kategori", "kebutuhan")
                sc(ws.cell(ri, 1, hari), align="center")
                sc(ws.cell(ri, 2, item["keterangan"]))
                sc(ws.cell(ri, 3, kat.capitalize()), align="center", bg=kc.get(kat, "FFFFFF"))
                sc(ws.cell(ri, 4, item["jumlah"]), align="right", num_fmt="#,##0", bg="FADBD8")
                total_keluar += item["jumlah"]
                kat_total[kat] += item["jumlah"]
                ri += 1

        sc(ws.cell(ri, 3, "TOTAL PENGELUARAN"), align="right")
        sc(ws.cell(ri, 4, total_keluar), align="right", num_fmt="#,##0", bg="FADBD8")
        ws.cell(ri, 3).font = Font(bold=True)
        ri += 2

        # Ringkasan bulan
        saldo_awal = data.get("saldo_awal") or 0
        saldo_akhir = saldo_awal + total_masuk - total_keluar
        total_ew = sum(info.get("saldo", 0) for info in data_ewallet.values())

        ws.cell(ri, 1, "RINGKASAN BULAN").font = Font(bold=True, color="FFFFFF")
        ws.cell(ri, 1).fill = hf("8E44AD")
        ws.merge_cells(f"A{ri}:D{ri}")
        ri += 1

        ringkasan_data = [
            ("Saldo Awal (Saku)", saldo_awal),
            ("Pemasukan Saku", total_masuk),
            ("Pengeluaran Saku", total_keluar),
            ("Saldo Saku Akhir", saldo_akhir),
            ("Total Saldo E-Wallet", total_ew),
            ("SALDO TOTAL", saldo_akhir + total_ew),
            ("Saldo Tabungan", data_tabungan.get("saldo", 0)),
        ]
        for label, val in ringkasan_data:
            sc(ws.cell(ri, 1, label))
            sc(ws.cell(ri, 2, val), align="right", num_fmt="#,##0",
               bg="D5F5E3" if val >= 0 else "FADBD8")
            if label in ("SALDO TOTAL", "Saldo Saku Akhir"):
                ws.cell(ri, 1).font = Font(bold=True)
                ws.cell(ri, 2).font = Font(bold=True)
            ri += 1

    # ── Sheet Ringkasan Semua Bulan ──
    ws2 = wb.create_sheet("Ringkasan Semua Bulan")
    ws2.sheet_view.showGridLines = False
    hdrs3 = ["Bulan", "Saldo Awal", "Pemasukan", "Pengeluaran", "Saldo Saku Akhir",
             "Total E-Wallet", "Saldo Total", "Status"]
    wdths3 = [14, 18, 18, 18, 20, 18, 18, 14]
    for col, (h, w) in enumerate(zip(hdrs3, wdths3), 1):
        sh(ws2.cell(1, col, h), bg="1A5276")
        ws2.column_dimensions[get_column_letter(col)].width = w

    total_ew = sum(info.get("saldo", 0) for info in data_ewallet.values())
    summary_rows = []
    for ri2, (bulan, data) in enumerate(data_keuangan.items(), 2):
        saldo_awal  = data.get("saldo_awal") or 0
        masuk  = sum(i["jumlah"] for h in data.get("pemasukan", {}).values()  for i in h)
        keluar = sum(i["jumlah"] for h in data.get("pengeluaran", {}).values() for i in h)
        akhir  = saldo_awal + masuk - keluar
        total  = akhir + total_ew
        pct    = (akhir / saldo_awal * 100) if saldo_awal > 0 else 0
        status = "Aman" if pct >= 70 else "Waspada" if pct >= 40 else "Hampir Habis" if pct >= 10 else "Defisit"
        bg_st  = "D5F5E3" if status == "Aman" else "FDEBD0" if status == "Waspada" else "FADBD8"
        for ci, val in enumerate([bulan, saldo_awal, masuk, keluar, akhir, total_ew, total, status], 1):
            sc(ws2.cell(ri2, ci, val), align="center",
               num_fmt="#,##0" if ci in [2, 3, 4, 5, 6, 7] else None,
               bg=bg_st if ci == 8 else ("F2F3F4" if ri2 % 2 == 0 else "FFFFFF"))
        summary_rows.append((bulan, masuk, keluar))

    if summary_rows:
        chart = BarChart(); chart.type = "col"; chart.style = 10
        chart.width = 22; chart.height = 14
        chart.title = "Pemasukan vs Pengeluaran per Bulan"
        chart.add_data(Reference(ws2, min_col=3, max_col=4, min_row=1, max_row=len(summary_rows)+1), titles_from_data=True)
        chart.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=len(summary_rows)+1))
        chart.series[0].graphicalProperties.solidFill = "27AE60"
        chart.series[1].graphicalProperties.solidFill = "E74C3C"
        ws2.add_chart(chart, "J2")

    # ── Sheet E-Wallet ──
    ws3 = wb.create_sheet("E-Wallet Tracker")
    ws3.sheet_view.showGridLines = False
    hdrs4 = ["E-Wallet", "Pengguna", "Bulan", "Jenis", "Kategori", "Keterangan", "Jumlah (Rp)"]
    wdths4 = [20, 14, 14, 14, 14, 35, 18]
    for col, (h, w) in enumerate(zip(hdrs4, wdths4), 1):
        sh(ws3.cell(1, col, h), bg="6C3483")
        ws3.column_dimensions[get_column_letter(col)].width = w
    kc2 = {"kebutuhan": "D5F5E3", "impulsif": "FADBD8", "tabungan": "D6EAF8"}
    ri3 = 2
    for nama, info in data_ewallet.items():
        for t in info.get("transaksi", []):
            row = [nama, info.get("pengguna", "-"), t.get("bulan", "-"),
                   t.get("jenis", "-").capitalize(), t["kategori"].capitalize(),
                   t["keterangan"], t["jumlah"]]
            for ci, val in enumerate(row, 1):
                sc(ws3.cell(ri3, ci, val), align="center" if ci < 6 else "left",
                   num_fmt="#,##0" if ci == 7 else None,
                   bg=kc2.get(t["kategori"], "FFFFFF"))
            ri3 += 1

    # ── Sheet Tabungan ──
    ws4 = wb.create_sheet("Tabungan")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "SALDO TABUNGAN"
    ws4["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws4["A1"].fill = hf("2980B9")
    ws4["A1"].alignment = Alignment(horizontal="center")
    ws4.merge_cells("A1:D1")
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 20

    sh(ws4.cell(2, 1, "Jenis"), bg="2471A3")
    sh(ws4.cell(2, 2, "Jumlah (Rp)"), bg="2471A3")
    sh(ws4.cell(2, 3, "Keterangan"), bg="2471A3")

    ri4 = 3
    for r in data_tabungan.get("riwayat", []):
        sc(ws4.cell(ri4, 1, r.get("jenis", "-").capitalize()), align="center",
           bg="D5F5E3" if r.get("jenis") == "tambah" else "FADBD8")
        sc(ws4.cell(ri4, 2, r.get("jumlah", 0)), align="right", num_fmt="#,##0")
        sc(ws4.cell(ri4, 3, r.get("keterangan", "-")))
        ri4 += 1

    sc(ws4.cell(ri4, 1, "SALDO TABUNGAN SAAT INI"))
    sc(ws4.cell(ri4, 2, data_tabungan.get("saldo", 0)), align="right", num_fmt="#,##0", bg="D6EAF8")
    ws4.cell(ri4, 1).font = Font(bold=True)
    ws4.cell(ri4, 2).font = Font(bold=True)

    path = os.path.join(os.getcwd(), "laporan_edubank.xlsx")
    wb.save(path)
    print(f"\n✅ Export berhasil! → laporan_edubank.xlsx")
    print("   Sheet: Data per Bulan | Ringkasan Semua Bulan | E-Wallet | Tabungan")

# ============================================================
# MENU 8 — E-WALLET: TAMBAH / UPDATE SALDO
# ============================================================
def input_saldo_ewallet():
    print("\n=== TAMBAH / UPDATE SALDO E-WALLET ===")
    daftar_default = ["GoPay", "ShopeePay", "DANA", "OVO", "LinkAja"]
    for i, nama in enumerate(daftar_default, 1):
        print(f"  {i}. {nama}")
    print(f"  {len(daftar_default)+1}. Lainnya (ketik sendiri)")

    try:
        pilihan = int(input("\nPilih e-wallet: "))
        if pilihan == len(daftar_default) + 1:
            nama_ew = input("Nama e-wallet: ").strip().title()
            if not nama_ew:
                print("❌ Nama tidak boleh kosong!"); return
        elif 1 <= pilihan <= len(daftar_default):
            nama_ew = daftar_default[pilihan - 1]
        else:
            print("❌ Pilihan tidak valid!"); return

        pengguna = input(f"Nama pengguna akun {nama_ew}: ").strip() or "Tidak diisi"
        saldo    = int(input(f"Saldo {nama_ew} saat ini: Rp "))
        key      = f"{nama_ew} ({pengguna})"

        if key not in data_ewallet:
            data_ewallet[key] = {"saldo": 0, "pengguna": pengguna, "transaksi": []}

        data_ewallet[key]["saldo"]    = saldo
        data_ewallet[key]["pengguna"] = pengguna
        print(f"\n✅ {key} → saldo Rp {saldo:,} berhasil disimpan!")
        print(f"   (Saldo e-wallet otomatis terakumulasi ke Saldo Total)")
    except:
        print("❌ Input tidak valid!")

# ============================================================
# MENU 9 — E-WALLET: CATAT TRANSAKSI
# ============================================================
def catat_transaksi_ewallet():
    if not data_ewallet:
        print("\n⚠️  Tambah e-wallet dulu! (Menu 8)"); return

    print("\n=== CATAT TRANSAKSI E-WALLET ===")

    print("Pilih bulan transaksi:")
    for b in DAFTAR_BULAN:
        print(f"  {b.capitalize()}", end="  ")
    print()
    bulan_ew = input("Ketik bulan: ").strip().lower()
    if bulan_ew not in DAFTAR_BULAN:
        print("❌ Bulan tidak valid!"); return
    bulan_ew = bulan_ew.capitalize()

    daftar = list(data_ewallet.keys())
    for i, nama in enumerate(daftar, 1):
        print(f"  {i}. {nama} (Saldo: Rp {data_ewallet[nama]['saldo']:,})")
    try:
        idx     = int(input("Pilih e-wallet: ")) - 1
        nama_ew = daftar[idx]
    except:
        print("❌ Input tidak valid!"); return

    keterangan = input("Keterangan transaksi: ").strip()
    if not keterangan:
        print("❌ Keterangan tidak boleh kosong!"); return

    print("\nJenis transaksi:")
    print("  1. Pengeluaran (uang keluar dari e-wallet)")
    print("  2. Pemasukan   (top up, transfer masuk)")
    jenis_pilih = input("Pilih (1/2): ").strip()

    if jenis_pilih == "1":
        jenis_ew = "pengeluaran"
        print("\nKategori pengeluaran:")
        print("  1. Kebutuhan  (makan, transport, bayar tagihan)")
        print("  2. Impulsif   (flash sale, jajan random, hiburan)")
        print("  3. Tabungan   (label saja, tidak nyambung menu Tabungan)")
        try:
            kat     = int(input("Pilih kategori (1/2/3): "))
            kat_map = {1: "kebutuhan", 2: "impulsif", 3: "tabungan"}
            if kat not in kat_map:
                print("❌ Tidak valid!"); return
            kategori = kat_map[kat]
        except:
            print("❌ Input tidak valid!"); return

    elif jenis_pilih == "2":
        jenis_ew = "pemasukan"
        kategori = "pemasukan"
    else:
        print("❌ Pilihan tidak valid!"); return

    try:
        jumlah = int(input("Jumlah: Rp "))
        if jumlah <= 0:
            print("❌ Jumlah harus lebih dari 0!"); return
    except:
        print("❌ Input harus angka!"); return

    if jenis_ew == "pengeluaran":
        data_ewallet[nama_ew]["saldo"] -= jumlah
    else:
        data_ewallet[nama_ew]["saldo"] += jumlah

    data_ewallet[nama_ew]["transaksi"].append({
        "bulan":      bulan_ew,
        "jenis":      jenis_ew,
        "kategori":   kategori,
        "keterangan": keterangan,
        "jumlah":     jumlah
    })

    print(f"\n✅ Transaksi {jenis_ew} Rp {jumlah:,} berhasil dicatat!")
    print(f"   Sisa saldo {nama_ew}: Rp {data_ewallet[nama_ew]['saldo']:,}")

# ============================================================
# MENU 10 — RINGKASAN E-WALLET
# ============================================================
def ringkasan_ewallet():
    if not data_ewallet:
        print("\n⚠️  Belum ada data e-wallet!"); return

    print(f"\n{'='*44}")
    print("  RINGKASAN E-WALLET")
    print(f"{'='*44}")

    total_saldo = 0
    total_imp   = 0
    total_keluar= 0

    for nama, info in data_ewallet.items():
        trx       = info.get("transaksi", [])
        keb       = sum(t["jumlah"] for t in trx if t["kategori"] == "kebutuhan")
        imp       = sum(t["jumlah"] for t in trx if t["kategori"] == "impulsif")
        tab       = sum(t["jumlah"] for t in trx if t["kategori"] == "tabungan")
        masuk_ew  = sum(t["jumlah"] for t in trx if t["jenis"] == "pemasukan")
        keluar_ew = keb + imp + tab

        print(f"\n  💳 {nama}")
        print(f"     Pengguna       : {info.get('pengguna', '-')}")
        print(f"     Saldo saat ini : Rp {info['saldo']:>10,}")
        print(f"     Pemasukan      : Rp {masuk_ew:>10,}")
        print(f"     Kebutuhan      : Rp {keb:>10,}")
        print(f"     Impulsif       : Rp {imp:>10,}")
        print(f"     Tabungan (lbl) : Rp {tab:>10,}")

        total_saldo   += info["saldo"]
        total_imp     += imp
        total_keluar  += keluar_ew

    # Hitung saldo total (saku + e-wallet)
    saldo_saku_akhir = hitung_saldo_akhir(bulan_aktif) if bulan_aktif else 0

    print(f"\n  {'─'*40}")
    print(f"  Total saldo semua e-wallet : Rp {total_saldo:,}")
    print(f"  Saldo saku ({bulan_aktif or '-'})     : Rp {saldo_saku_akhir:,}")
    print(f"  SALDO TOTAL                : Rp {total_saldo + saldo_saku_akhir:,}")

    if total_keluar > 0:
        pct = total_imp / total_keluar * 100
        print(f"  % Pengeluaran impulsif     : {pct:.1f}%")
        if pct >= 50:
            print("\n  ⚠️  PERINGATAN: Lebih dari setengah uang e-wallet habis impulsif!")
        elif pct >= 30:
            print("\n  ⚠️  WASPADA: Pengeluaran impulsif cukup tinggi.")
        else:
            print("\n  ✅ Bagus! Pengeluaran e-wallet kamu terkontrol.")
    print(f"{'='*44}")

# ============================================================
# MENU 11 — TABUNGAN: TAMBAH / KURANGI SALDO
# ============================================================
def kelola_tabungan():
    print(f"\n=== MENU TABUNGAN ===")
    print(f"  Saldo Tabungan Saat Ini: Rp {data_tabungan['saldo']:,}")
    print("\n  1. Tambah saldo tabungan")
    print("  2. Kurangi saldo tabungan")
    print("  3. Lihat riwayat tabungan")
    print("  4. Kembali")

    pilihan = input("Pilih: ").strip()
    if pilihan == "1":
        try:
            jml = int(input("Jumlah tambah: Rp "))
            ket = input("Keterangan: ").strip() or "Menabung"
            data_tabungan["saldo"] += jml
            data_tabungan["riwayat"].append({"jenis": "tambah", "jumlah": jml, "keterangan": ket})
            print(f"✅ Saldo tabungan bertambah Rp {jml:,}. Total: Rp {data_tabungan['saldo']:,}")
        except:
            print("❌ Input tidak valid!")
    elif pilihan == "2":
        try:
            jml = int(input("Jumlah kurangi: Rp "))
            if jml > data_tabungan["saldo"]:
                print("❌ Saldo tabungan tidak cukup!"); return
            ket = input("Keterangan: ").strip() or "Pengambilan tabungan"
            data_tabungan["saldo"] -= jml
            data_tabungan["riwayat"].append({"jenis": "kurang", "jumlah": jml, "keterangan": ket})
            print(f"✅ Saldo tabungan berkurang Rp {jml:,}. Sisa: Rp {data_tabungan['saldo']:,}")
        except:
            print("❌ Input tidak valid!")
    elif pilihan == "3":
        if not data_tabungan["riwayat"]:
            print("Belum ada riwayat tabungan.")
        else:
            print(f"\n{'─'*44}")
            for r in data_tabungan["riwayat"]:
                simbol = "+" if r["jenis"] == "tambah" else "-"
                print(f"  {simbol} Rp {r['jumlah']:>10,}  |  {r['keterangan']}")
            print(f"{'─'*44}")
            print(f"  TOTAL SALDO TABUNGAN: Rp {data_tabungan['saldo']:,}")

# ============================================================
# MENU 12 — SIMULASI NABUNG
# ============================================================
def simulasi_nabung():
    print("\n=== SIMULASI MENABUNG ===")
    print(f"  Saldo Tabungan Saat Ini: Rp {data_tabungan['saldo']:,}")
    try:
        target    = int(input("Target tabungan      : Rp "))
        per_bulan = int(input("Sanggup sisihkan/bln : Rp "))
        if per_bulan <= 0:
            print("❌ Harus lebih dari 0!"); return

        saldo_skrng = data_tabungan["saldo"]
        sisa_target = max(0, target - saldo_skrng)

        if sisa_target == 0:
            print(f"\n  🎉 Target sudah tercapai! Saldo tabunganmu Rp {saldo_skrng:,}")
            return

        bulan_butuh = (sisa_target + per_bulan - 1) // per_bulan
        sisa        = sisa_target % per_bulan

        print(f"\n  Target            : Rp {target:,}")
        print(f"  Saldo saat ini    : Rp {saldo_skrng:,}")
        print(f"  Sisa yg dibutuhkan: Rp {sisa_target:,}")
        print(f"  Per bulan         : Rp {per_bulan:,}")
        print(f"  Estimasi          : {bulan_butuh} bulan{' + sedikit' if sisa else ''}")

        if bulan_butuh <= 3:    print("  🚀 Sangat realistis! Semangat!")
        elif bulan_butuh <= 6:  print("  💪 Butuh konsisten, tapi bisa!")
        elif bulan_butuh <= 12: print("  📅 Butuh hampir setahun. Naikkan sisihan kalau bisa.")
        else:                   print("  📈 Cukup lama. Pertimbangkan naikkan jumlah sisihan.")

        print(f"\n  Progres tabungan:")
        terkumpul = saldo_skrng
        for i in range(1, bulan_butuh + 2):
            terkumpul = min(terkumpul + per_bulan, target)
            if i % 3 == 0 or i == 1 or terkumpul >= target:
                pct = terkumpul / target * 100
                bar = "█" * int(pct // 10) + "░" * (10 - int(pct // 10))
                print(f"  Bulan {i:>2} [{bar}] {pct:>5.1f}% — Rp {terkumpul:,}")
            if terkumpul >= target:
                break
    except:
        print("❌ Input tidak valid!")

# ============================================================
# MENU 13 — GRAFIK MATPLOTLIB
# ============================================================
def tampilkan_grafik():
    import matplotlib.pyplot as plt
    import numpy as np

    fig, axes = plt.subplots(1, 3, figsize=(18, 6))
    fig.suptitle("📊 Dashboard EduBank", fontsize=15, fontweight="bold")
    fig.patch.set_facecolor("#F8F9FA")

    # Chart 1: Pie pengeluaran per hari (bulan aktif)
    ax1 = axes[0]; ax1.set_facecolor("#F8F9FA")
    if bulan_aktif and bulan_aktif in data_keuangan:
        labels, values = [], []
        for hari, items in data_keuangan[bulan_aktif]["pengeluaran"].items():
            t = sum(i["jumlah"] for i in items)
            if t > 0: labels.append(hari); values.append(t)
        if values:
            ax1.pie(values, labels=labels, autopct="%1.1f%%", startangle=90,
                    colors=plt.cm.Set3.colors[:len(values)],
                    wedgeprops={"edgecolor": "white", "linewidth": 2})
        else:
            ax1.text(0.5, 0.5, "Belum ada data", ha="center", va="center", transform=ax1.transAxes)
    ax1.set_title(f"Pengeluaran/Hari ({bulan_aktif})", fontsize=12, fontweight="bold")

    # Chart 2: Bar kategori gabungan (saku + e-wallet)
    ax2 = axes[1]; ax2.set_facecolor("#F8F9FA")
    kat_total = {"Kebutuhan": 0, "Impulsif": 0, "Tabungan": 0}
    if bulan_aktif and bulan_aktif in data_keuangan:
        for items in data_keuangan[bulan_aktif]["pengeluaran"].values():
            for item in items:
                kat = item.get("kategori", "kebutuhan").capitalize()
                if kat in kat_total: kat_total[kat] += item["jumlah"]
    for info in data_ewallet.values():
        for t in info.get("transaksi", []):
            if t.get("jenis") == "pengeluaran":
                kat = t.get("kategori", "kebutuhan").capitalize()
                if kat in kat_total: kat_total[kat] += t["jumlah"]
    colors = ["#27AE60", "#E74C3C", "#2980B9"]
    ax2.bar(list(kat_total.keys()), list(kat_total.values()), color=colors, edgecolor="white")
    ax2.set_title("Kategori Pengeluaran\n(Saku + E-Wallet)", fontsize=12, fontweight="bold")
    ax2.spines["top"].set_visible(False); ax2.spines["right"].set_visible(False)

    # Chart 3: Bar e-wallet per dompet
    ax3 = axes[2]; ax3.set_facecolor("#F8F9FA")
    if data_ewallet:
        names, keb, imp, tab = [], [], [], []
        for nama, info in data_ewallet.items():
            names.append(nama)
            keb.append(sum(t["jumlah"] for t in info["transaksi"] if t["kategori"] == "kebutuhan"))
            imp.append(sum(t["jumlah"] for t in info["transaksi"] if t["kategori"] == "impulsif"))
            tab.append(sum(t["jumlah"] for t in info["transaksi"] if t["kategori"] == "tabungan"))
        x = np.arange(len(names)); w = 0.25
        ax3.bar(x - w, keb, w, label="Kebutuhan", color="#27AE60", edgecolor="white")
        ax3.bar(x,     imp, w, label="Impulsif",  color="#E74C3C", edgecolor="white")
        ax3.bar(x + w, tab, w, label="Tabungan",  color="#2980B9", edgecolor="white")
        ax3.set_xticks(x); ax3.set_xticklabels(names, fontsize=9)
        ax3.legend(); ax3.spines["top"].set_visible(False); ax3.spines["right"].set_visible(False)
    else:
        ax3.text(0.5, 0.5, "Belum ada data e-wallet", ha="center", va="center", transform=ax3.transAxes)
    ax3.set_title("E-Wallet per Kategori", fontsize=12, fontweight="bold")

    plt.tight_layout(); plt.show()
    print("✅ Grafik ditampilkan!")

# ============================================================
# MENU UTAMA
# ============================================================
load_data()

while True:
    saldo_info = f"Bulan: {bulan_aktif}" if bulan_aktif else "Belum pilih bulan"
    saldo_akhir_info = f"Rp {hitung_saldo_akhir(bulan_aktif):,}" if bulan_aktif and bulan_aktif in data_keuangan else "-"
    total_ew_info = f"Rp {sum(i.get('saldo',0) for i in data_ewallet.values()):,}"
    tab_info = f"Rp {data_tabungan['saldo']:,}"

    print(f"╔═════════════════════════════════════╗")
    print(f"║             EDUBANK                 ║")
    print(f"║  {saldo_info:<32}                   ║")
    print(f"║ Saldo Saku  : {saldo_akhir_info:<18}║")
    print(f"║ Saldo E-Wal : {total_ew_info:<18}   ║")
    print(f"║ Tabungan    : {tab_info:<18}        ║")
    print(f"╠═════════════════════════════════════╣")
    print(f"║  PENCATATAN KEUANGAN (SAKU)         ║")
    print(f"║  1.  Pilih Bulan                    ║")
    print(f"║  2.  Input Saldo Awal Saku          ║")
    print(f"║  3.  Tambah Transaksi Saku          ║")
    print(f"║  4.  Ringkasan Bulan Ini            ║")
    print(f"║  5.  Total Per Hari                 ║")
    print(f"║  6.  Edit / Hapus Transaksi         ║")
    print(f"║  7.  Export Excel                   ║")
    print(f"╠═════════════════════════════════════╣")
    print(f"║  E-WALLET TRACKER                   ║")
    print(f"║  8.  Tambah / Update Saldo EW       ║")
    print(f"║  9.  Catat Transaksi E-Wallet       ║")
    print(f"║  10. Ringkasan E-Wallet             ║")
    print(f"╠═════════════════════════════════════╣")
    print(f"║  TABUNGAN                           ║")
    print(f"║  11. Kelola Saldo Tabungan          ║")
    print(f"║  12. Simulasi Menabung              ║")
    print(f"╠═════════════════════════════════════╣")
    print(f"║  13. Tampilkan Grafik               ║")
    print(f"║  14. Simpan & Keluar                ║")
    print(f"╚═════════════════════════════════════╝")

    p = input("Pilih menu: ").strip()

    if   p == "1":  pilih_bulan()
    elif p == "2":  input_saldo()
    elif p == "3":  tambah_transaksi()
    elif p == "4":  ringkasan()
    elif p == "5":  total_per_day()
    elif p == "6":  edit_hapus()
    elif p == "7":  export_excel()
    elif p == "8":  input_saldo_ewallet()
    elif p == "9":  catat_transaksi_ewallet()
    elif p == "10": ringkasan_ewallet()
    elif p == "11": kelola_tabungan()
    elif p == "12": simulasi_nabung()
    elif p == "13": tampilkan_grafik()
    elif p == "14":
        simpan_data()
        print(" Terima kasih sudah pakai EduBank! 💰")
        break
    else:
        print("❌ Pilihan tidak valid.")
